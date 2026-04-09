"""
Ewidencja Przebiegu Pojazdu – Generator
Smart Work Sp. z o.o.
v4.0 – wbudowany czytnik XLS (bez LibreOffice!)
"""

import os, io, datetime, calendar, re, struct, json, time
from flask import Flask, request, send_file, jsonify, make_response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    import requests
except ImportError:
    requests = None

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024


# ══════════════════════════════════════════════════════════════════════════════
# WBUDOWANY CZYTNIK XLS  (OLE + BIFF8) – bez LibreOffice, bez xlrd
# ══════════════════════════════════════════════════════════════════════════════

def _xls_read_rows(xls_bytes, date_cols=(8, 11)):
    """
    Czyta plik .xls (BIFF8/OLE) z czystego Pythona.
    Zwraca listę wierszy (list of lists).
    date_cols – numery kolumn traktowane jako daty/godziny.
    """
    MAGIC       = b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'
    ENDOFCHAIN  = 0xFFFFFFFE
    FREESECT    = 0xFFFFFFFF

    if xls_bytes[:8] != MAGIC:
        raise ValueError("Plik nie jest formatem XLS (OLE). Użyj .xls z GPS trackera.")

    data = xls_bytes

    # ── OLE header ────────────────────────────────────────────────────────────
    sec_size      = 2 ** struct.unpack_from('<H', data, 30)[0]
    mini_sec_size = 2 ** struct.unpack_from('<H', data, 32)[0]
    first_dir     = struct.unpack_from('<I', data, 48)[0]
    mini_cutoff   = struct.unpack_from('<I', data, 56)[0]
    first_minifat = struct.unpack_from('<I', data, 60)[0]

    def get_sector(sec):
        off = 512 + sec * sec_size
        return data[off: off + sec_size]

    # Build FAT
    fat = []
    for i in range(109):
        sec = struct.unpack_from('<I', data, 76 + i * 4)[0]
        if sec in (FREESECT, ENDOFCHAIN): break
        chunk = get_sector(sec)
        for j in range(0, len(chunk), 4):
            fat.append(struct.unpack_from('<I', chunk, j)[0])

    def follow(start):
        chain, sec = [], start
        while sec not in (ENDOFCHAIN, FREESECT) and sec < len(fat):
            chain.append(sec); sec = fat[sec]
        return chain

    # Directory
    dir_data = b''.join(get_sector(s) for s in follow(first_dir))
    entries  = {}
    root_entry = None
    for i in range(len(dir_data) // 128):
        e     = dir_data[i * 128:(i + 1) * 128]
        nlen  = struct.unpack_from('<H', e, 64)[0]
        name  = e[:max(nlen - 2, 0)].decode('utf-16-le', errors='replace') if nlen >= 2 else ''
        etype = e[66]
        start = struct.unpack_from('<I', e, 116)[0]
        size  = struct.unpack_from('<I', e, 120)[0]
        entries[i] = dict(name=name, type=etype, start=start, size=size)
        if i == 0: root_entry = entries[i]

    # Find Workbook/Book stream
    wb_entry = next(
        (e for e in entries.values()
         if e['name'].lower() in ('workbook', 'book') and e['type'] == 2),
        None
    )
    if wb_entry is None:
        raise ValueError("Nie znaleziono strumienia Workbook w pliku XLS.")

    # Read stream (normal or mini)
    if wb_entry['size'] < mini_cutoff and root_entry:
        mini_data = b''.join(get_sector(s) for s in follow(root_entry['start']))
        minifat   = []
        for s in follow(first_minifat):
            chunk = get_sector(s)
            for j in range(0, len(chunk), 4):
                minifat.append(struct.unpack_from('<I', chunk, j)[0])
        result, sec = b'', wb_entry['start']
        while sec not in (ENDOFCHAIN, FREESECT) and sec < len(minifat):
            off = sec * mini_sec_size
            result += mini_data[off: off + mini_sec_size]
            sec = minifat[sec]
        stream = result[:wb_entry['size']]
    else:
        stream = b''.join(get_sector(s) for s in follow(wb_entry['start']))
        stream = stream[:wb_entry['size']]

    # ── BIFF8 record parser ───────────────────────────────────────────────────
    XL_EPOCH = datetime.datetime(1899, 12, 30)

    def xl_dt(v):
        try:
            days = int(v)
            secs = round((v - days) * 86400)
            return XL_EPOCH + datetime.timedelta(days=days, seconds=secs)
        except Exception:
            return v

    cells = {}
    sst   = []
    pos   = 0
    sd    = stream

    # We also collect extra SST data from CONTINUE records
    sst_raw = b''
    in_sst  = False

    while pos + 4 <= len(sd):
        rt  = struct.unpack_from('<H', sd, pos)[0]
        rl  = struct.unpack_from('<H', sd, pos + 2)[0]
        rec = sd[pos + 4: pos + 4 + rl]
        pos += 4 + rl

        # ── SST ──────────────────────────────────────────────────────────────
        if rt == 0x00FC:
            in_sst   = True
            sst_raw  = rec
            sst_full = sst_raw
            # Parse SST
            total = struct.unpack_from('<I', sst_full, 0)[0]
            count = struct.unpack_from('<I', sst_full, 4)[0]
            p2    = 8
            for _ in range(count):
                if p2 + 3 > len(sst_full): break
                cch   = struct.unpack_from('<H', sst_full, p2)[0]
                flags = sst_full[p2 + 2]
                p2   += 3
                comp  = not (flags & 1)
                if flags & 8 and p2 + 2 <= len(sst_full):
                    rich = struct.unpack_from('<H', sst_full, p2)[0]; p2 += 2
                else:
                    rich = 0
                if flags & 4 and p2 + 4 <= len(sst_full):
                    p2 += 4
                blen = cch * (1 if comp else 2)
                try:
                    s = sst_full[p2:p2 + blen].decode('latin-1' if comp else 'utf-16-le', errors='replace')
                except Exception:
                    s = ''
                p2  += blen + rich * 4
                sst.append(s)
            in_sst = False

        # ── LABELSST ─────────────────────────────────────────────────────────
        elif rt == 0x00FD and len(rec) >= 8:
            r, c  = struct.unpack_from('<HH', rec, 0)
            idx   = struct.unpack_from('<I',  rec, 6)[0]
            cells[(r, c)] = sst[idx] if idx < len(sst) else ''

        # ── LABEL ─────────────────────────────────────────────────────────────
        elif rt == 0x0204 and len(rec) >= 9:
            r, c  = struct.unpack_from('<HH', rec, 0)
            cch   = struct.unpack_from('<H',  rec, 6)[0]
            flags = rec[8]
            p2    = 9
            blen  = cch * (1 if not (flags & 1) else 2)
            try:
                s = rec[p2:p2 + blen].decode('latin-1' if not (flags & 1) else 'utf-16-le', errors='replace')
            except Exception:
                s = ''
            cells[(r, c)] = s.strip()

        # ── NUMBER ────────────────────────────────────────────────────────────
        elif rt == 0x0203 and len(rec) >= 14:
            r, c = struct.unpack_from('<HH', rec, 0)
            val  = struct.unpack_from('<d',  rec, 6)[0]
            cells[(r, c)] = xl_dt(val) if c in date_cols and 1 < val < 3e5 else val

        # ── RK ───────────────────────────────────────────────────────────────
        elif rt == 0x027E and len(rec) >= 10:
            r, c = struct.unpack_from('<HH', rec, 0)
            rk   = struct.unpack_from('<I',  rec, 6)[0]
            if rk & 2:
                val = (rk >> 2) / (100.0 if rk & 1 else 1.0)
            else:
                val = struct.unpack('<d', bytes(4) + struct.pack('<I', rk & 0xFFFFFFFC))[0]
                if rk & 1: val /= 100.0
            cells[(r, c)] = xl_dt(val) if c in date_cols and 1 < val < 3e5 else val

        # ── MULRK ─────────────────────────────────────────────────────────────
        elif rt == 0x00BD and len(rec) >= 6:
            r    = struct.unpack_from('<H', rec, 0)[0]
            fcol = struct.unpack_from('<H', rec, 2)[0]
            n    = (len(rec) - 2) // 6
            for i in range(n):
                c  = fcol + i
                rk = struct.unpack_from('<I', rec, 4 + i * 6 + 2)[0]
                if rk & 2:
                    val = (rk >> 2) / (100.0 if rk & 1 else 1.0)
                else:
                    val = struct.unpack('<d', bytes(4) + struct.pack('<I', rk & 0xFFFFFFFC))[0]
                    if rk & 1: val /= 100.0
                cells[(r, c)] = xl_dt(val) if c in date_cols and 1 < val < 3e5 else val

    if not cells:
        return []
    max_r = max(r for r, c in cells) + 1
    max_c = max(c for r, c in cells) + 1
    return [[cells.get((r, co)) for co in range(max_c)] for r in range(max_r)]


# ══════════════════════════════════════════════════════════════════════════════
# GPS PARSING
# ══════════════════════════════════════════════════════════════════════════════

def _remove_digits_from_name(text):
    """Strip house numbers, zip codes, and common street prefixes."""
    if not text:
        return ''
    text = str(text)
    # Remove Polish/German zip codes like 43-245
    text = re.sub(r'\b\d{2}-\d{3}\b', '', text)
    # Remove standalone numbers (house numbers like 32, 44)
    text = re.sub(r'\b\d+\b', '', text)
    # Remove street prefixes
    text = re.sub(r'\b(ulica|ul\.|al\.|aleja|plac|pl\.)\s+', '', text, flags=re.IGNORECASE)
    # Remove administrative divisions: Powiat, Gmina, Województwo (incl. abbreviations pow., gm., woj.)
    text = re.sub(r'\b(Powiat|Gmina|Województwo|gm\.|pow\.|woj\.)\s+.*$', '', text, flags=re.IGNORECASE)
    # Collapse whitespace and strip trailing punctuation
    return ' '.join(text.split()).strip(' ,.')


_STREET_PREFIX_RE = re.compile(
    r'\b(ulica|ul\.|al\.|aleja|plac|pl\.|straße|str\.'
    r'|avenue|ave\.|road|rd\.|street|st\.)\b',
    re.IGNORECASE
)

def _looks_like_street(text):
    """
    Returns True when the text segment looks like a real street address
    (contains a digit – house number – or starts with a known street prefix).
    POI / business names like "stop cafe" or "Centrum Handlowe" return False.
    """
    if not text:
        return False
    if re.search(r'\d', text):          # has any digit → street with house number
        return True
    if _STREET_PREFIX_RE.search(text):  # starts with ul., ulica, al. …
        return True
    return False

# Default hardcoded cities if user doesn't use the JSON file
ZIP_TO_CITY = {
    "43-245": "Studzionka",
    "43-246": "Strumień",
}

KNOWN_CITIES_FILE = 'known_cities.json'

def load_known_cities():
    if os.path.exists(KNOWN_CITIES_FILE):
        try:
            with open(KNOWN_CITIES_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return {**ZIP_TO_CITY, **data}
        except Exception:
            pass
    return ZIP_TO_CITY.copy()

def save_known_cities(cities_dict):
    try:
        # Don't save the hardcoded ones if they weren't modified
        to_save = {k: v for k, v in cities_dict.items() if k not in ZIP_TO_CITY or ZIP_TO_CITY[k] != v}
        with open(KNOWN_CITIES_FILE, 'w', encoding='utf-8') as f:
            json.dump(to_save, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print(f"Error saving known cities: {e}")

_ONLINE_CITIES_CACHE = load_known_cities()

def geocode_city_nominatim(zip_code, street_raw):
    """Hits Nominatim to find the city. Returns city name or None."""
    if not requests:
        return None
        
    query = f"{zip_code} {street_raw}, Poland"
    url = "https://nominatim.openstreetmap.org/search"
    params = {
        'q': query,
        'format': 'jsonv2',
        'addressdetails': 1,
        'limit': 1
    }
    headers = {
        'User-Agent': 'GeneratorEwidencji/4.0 (contact@example.com)'
    }
    
    try:
        # Respect OpenStreetMap's 1 request/sec policy
        time.sleep(1.1)
        resp = requests.get(url, params=params, headers=headers, timeout=5)
        if resp.status_code == 200:
            data = resp.json()
            if data and len(data) > 0:
                address = data[0].get('address', {})
                # Try various keys where city names can be mapped
                city = (address.get('city') or 
                        address.get('town') or 
                        address.get('village') or 
                        address.get('municipality'))
                return city
    except Exception as e:
        print(f"Nominatim Error: {e}")
    return None

# Regex that matches a segment starting with a ZIP code followed by city name,
# e.g. "55-200 Oława" – GPS sometimes omits the comma between them.
_ZIP_CITY_RE = re.compile(r'^(\d{2}-\d{3})\s+(.+)$')


def _parse_gps_addr(addr):
    """
    Parse a raw GPS address string.
    GPS format: "Street HouseNo, ZipCode, CityName, Country"
              or "Street HouseNo, CityName, Country"
              or "Street HouseNo, ZipCode, Country"  (no city name!)
              or "POI Name, Street HouseNo, ZipCode City, Country"  ← POI first!

    Returns:
        street_clean  – street name without house number
        city_name     – actual city name (empty string if not available)
        city_key      – city identifier for grouping (city_name if available, else zip code)
    """
    if not addr:
        return '', '', ''

    raw = str(addr)
    # Remove country suffixes
    raw = raw.replace(', Poland', '').replace(', Polska', '')
    raw = raw.strip().rstrip(',')

    parts = [p.strip() for p in raw.split(',') if p.strip()]
    if not parts:
        return '', '', ''

    # ── POI detection ────────────────────────────────────────────────────────
    # If the first segment does NOT look like a street (no digits, no "ul." etc.)
    # but the second segment DOES look like a street → first segment is a POI name;
    # skip it and use the rest as-is.
    # Example: "stop cafe, ulica Opolska 31, 55-200 Oława"
    #           ↑ POI name     ↑ real street
    if (len(parts) >= 2
            and not _looks_like_street(parts[0])
            and _looks_like_street(parts[1])):
        parts = parts[1:]   # drop POI segment

    # Regex for a pure zip code part like "43-245"
    zip_re = re.compile(r'^\d{2}-\d{3}$')

    street_raw = parts[0]
    street_clean = _remove_digits_from_name(street_raw)

    # Collect non-street parts
    rest = parts[1:]
    zip_code = ''
    city_name = ''

    for part in rest:
        if zip_re.match(part):
            if not zip_code:
                zip_code = part   # remember first zip found
        else:
            # Check for combined "55-200 CityName" format (GPS sometimes merges them)
            m_zip_city = _ZIP_CITY_RE.match(part)
            if m_zip_city:
                if not zip_code:
                    zip_code = m_zip_city.group(1)
                # Prefer the city name associated with the zip code and stop looking further
                city_name = m_zip_city.group(2).strip()
                break
            else:
                # Only set city_name if we haven't found one yet
                if not city_name:
                    p = part.strip()
                    # Skip parts that look like administrative regions
                    if not re.match(r'^(Powiat|Gmina|Województwo|gm\.|pow\.|woj\.)', p, re.I):
                        city_name = p

    if not city_name and zip_code:
        # Check cache
        if zip_code in _ONLINE_CITIES_CACHE:
            city_name = _ONLINE_CITIES_CACHE[zip_code]
        else:
            # Try to fetch from API
            found_city = geocode_city_nominatim(zip_code, street_raw)
            if found_city:
                city_name = found_city
                _ONLINE_CITIES_CACHE[zip_code] = city_name
                save_known_cities(_ONLINE_CITIES_CACHE)

    # city_key is used internally for "is this a different city?" check
    # Use real city name if we have it, otherwise fall back to zip
    city_key = city_name if city_name else zip_code

    return street_clean, city_name, city_key


def shorten_addr(addr):
    """Return a display string for a single address: 'Street, City' or just 'Street'."""
    if not addr:
        return ''
    street, city_name, _ = _parse_gps_addr(addr)
    street = _remove_digits_from_name(street)
    city   = _remove_digits_from_name(city_name)
    if street and city and street.lower() != city.lower():
        return f'{street}, {city}'
    return street or city


def extract_city(addr):
    """Return only the city name (or city key) for multi-city detection."""
    if not addr:
        return ''
    _, city_name, city_key = _parse_gps_addr(addr)
    return _remove_digits_from_name(city_name.strip() if city_name else '')



def parse_gps(file_bytes, filename):
    ext = os.path.splitext(filename)[1].lower()

    if ext == '.xls':
        rows = _xls_read_rows(file_bytes)
    elif ext == '.xlsx':
        from openpyxl import load_workbook
        import io as _io
        wb = load_workbook(_io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    else:
        raise ValueError(f'Nieobsługiwany format pliku: {ext}. Użyj .xls lub .xlsx')

    plate = car_model = date_from = date_to = None

    for i, row in enumerate(rows[:20]):
        if not row: continue
        v = str(row[0]) if row[0] else ''
        if 'Data:' in v:
            m = re.search(r'(\d{4}-\d{2}-\d{2}).*?(\d{4}-\d{2}-\d{2})', v)
            if m:
                date_from = datetime.date.fromisoformat(m.group(1))
                date_to   = datetime.date.fromisoformat(m.group(2))
        if i == 12:
            car_model = str(row[0]).strip() if row[0] else ''
            plate     = str(row[2]).strip() if len(row) > 2 and row[2] else ''

    day_groups, current = [], []

    for row in rows:
        if not row or len(row) < 17:
            current.append(row); continue
        first = str(row[0]).strip() if row[0] else ''
        if first == 'Razem:':
            if current:
                data = [r for r in current
                        if len(r) > 12 and isinstance(r[8], datetime.datetime)]
                if data:
                    if not plate:
                        plate = str(data[0][0]).strip()
                    date = data[0][8].date()
                    km   = float(row[16]) if isinstance(row[16], (int, float)) else 0.0
                    addresses = []
                    for r in data:
                        s, e = r[9], r[12]
                        if s: addresses.append(str(s))
                        if e: addresses.append(str(e))
                        
                    first_start = data[0][9] if data[0][9] else ''
                    last_end = data[-1][12] if data[-1][12] else ''

                    day_groups.append(dict(date=date, km=km,
                                          addresses=addresses,
                                          first_start=first_start, last_end=last_end))
            current = []
        else:
            current.append(row)

    return plate, car_model, date_from, date_to, day_groups


def get_polish_holidays(year):
    """
    Returns a set of datetime.date objects representing Polish public holidays for a given year.
    Includes fixed dates and calculates Easter (and dependent holidays) using Computus.
    """
    holidays = set()
    
    # Fixed holidays
    holidays.add(datetime.date(year, 1, 1))   # Nowy Rok
    holidays.add(datetime.date(year, 1, 6))   # Trzech Króli
    holidays.add(datetime.date(year, 5, 1))   # Święto Pracy
    holidays.add(datetime.date(year, 5, 3))   # Święto Konstytucji 3 Maja
    holidays.add(datetime.date(year, 8, 15))  # Wniebowzięcie NMP
    holidays.add(datetime.date(year, 11, 1))  # Wszystkich Świętych
    holidays.add(datetime.date(year, 11, 11)) # Święto Niepodległości
    holidays.add(datetime.date(year, 12, 24)) # Wigilia Bożego Narodzenia (na życzenie)
    holidays.add(datetime.date(year, 12, 25)) # Boże Narodzenie (dzień 1)
    holidays.add(datetime.date(year, 12, 26)) # Boże Narodzenie (dzień 2)
    
    # Calculate Easter Sunday (Meeus/Jones/Butcher algorithm)
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    
    easter_sunday = datetime.date(year, month, day)
    
    # Dependent holidays
    holidays.add(easter_sunday)                                      # Wielkanoc
    holidays.add(easter_sunday + datetime.timedelta(days=1))         # Poniedziałek Wielkanocny
    holidays.add(easter_sunday + datetime.timedelta(days=60))        # Boże Ciało
    
    return holidays


def get_previous_working_day(target_date):
    """
    Returns the nearest previous working day (mon-fri) that is not a public holiday.
    If the previous working day falls in the previous month, 
    returns the next available working day in the target month instead.
    """
    current_date = target_date
    while True:
        holidays = get_polish_holidays(current_date.year)
        # Check if it's weekend (5=Sat, 6=Sun) or a public holiday
        if current_date.weekday() >= 5 or current_date in holidays:
            current_date -= datetime.timedelta(days=1)
        else:
            break
            
    if current_date.month == target_date.month and current_date.year == target_date.year:
        return current_date
        
    # If shifting backwards placed us in the previous month, go forwards instead
    current_date = target_date
    while True:
        holidays = get_polish_holidays(current_date.year)
        if current_date.weekday() >= 5 or current_date in holidays:
            current_date += datetime.timedelta(days=1)
        else:
            return current_date


def aggregate(day_groups):
    agg = {}
    for g in sorted(day_groups, key=lambda x: x['date']):
        d = g['date']
        # Calculate the proper working day for this log entry
        t = get_previous_working_day(d)
        
        if t not in agg:
            agg[t] = dict(km=0.0, addresses=[], first_loc=None, last_loc=None)
            
        agg[t]['km'] += g['km']
        agg[t]['addresses'].extend(g['addresses'])
        
        if 'daily_addresses' not in agg[t]:
            agg[t]['daily_addresses'] = []
        if g['addresses']:
            agg[t]['daily_addresses'].append(g['addresses'])
            
        # The first location logged for this working day
        if agg[t]['first_loc'] is None and g.get('first_start'):
            agg[t]['first_loc'] = g['first_start']
            
        # The last location logged simply overrides any previous ones
        if g.get('last_end'):
            agg[t]['last_loc'] = g['last_end']
            
    return agg


def aggregate_actual(day_groups):
    """
    Sumuje km per faktyczny dzień jazdy.
    NIE przesuwa weekendów/świąt – każdy dzień zostaje tam gdzie był.
    """
    agg = {}
    for g in sorted(day_groups, key=lambda x: x['date']):
        d = g['date']   # prawdziwa data, bez przesunięcia
        if d not in agg:
            agg[d] = dict(km=0.0, addresses=[], first_loc=None, last_loc=None)

        agg[d]['km'] += g['km']
        agg[d]['addresses'].extend(g['addresses'])

        if 'daily_addresses' not in agg[d]:
            agg[d]['daily_addresses'] = []
        if g['addresses']:
            agg[d]['daily_addresses'].append(g['addresses'])

        if agg[d]['first_loc'] is None and g.get('first_start'):
            agg[d]['first_loc'] = g['first_start']

        if g.get('last_end'):
            agg[d]['last_loc'] = g['last_end']

    return agg


def build_route(addresses, last_known_city=""):
    if not addresses:
        return '', last_known_city
        
    cities_seq = []
    streets_in_city = []
    unique_cities = []
    
    current_city = last_known_city
    
    for addr in addresses:
        street, city_name, city_key = _parse_gps_addr(addr)
        
        street = _remove_digits_from_name(street)
        if ',' in street:
            street = street.split(',')[0].strip()
            
        city = _remove_digits_from_name(city_name)
        
        if not city:
            # Fallback to the last known city if GPS lost the city name
            city = current_city
        else:
            current_city = city
            
        if city and city not in unique_cities:
            unique_cities.append(city)
            
        if city and (not cities_seq or cities_seq[-1] != city):
            cities_seq.append(city)
            
        if street and street not in streets_in_city:
            # Clean up potential duplicates if street is exactly the city name
            if street.lower() != city.lower():
                streets_in_city.append(street)
            
    valid_cities = [c for c in cities_seq if c]
    
    if len(unique_cities) > 1:
        # Multiple cities format: CityA-CityB-CityA
        return '-'.join(valid_cities), current_city
    else:
        # Single city format: CityA: Street1, Street2
        city = unique_cities[0] if unique_cities else current_city
        streets = [s for s in streets_in_city if s and s.lower() != city.lower()]
        
        if city and streets:
            return f"{city}: {', '.join(streets)}", current_city
        elif city:
            return city, current_city
        else:
            return ', '.join(streets), current_city


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL GENERATION
# ══════════════════════════════════════════════════════════════════════════════

PL_DAYS = {0:'pon',1:'wt',2:'śr',3:'czw',4:'pt',5:'sob',6:'niedz'}
DARK, ACCENT, GRAY = '1F2937', '10B981', '6B7280'


def bdr(top='thin', right='thin', bottom='thin', left='thin'):
    def s(v): return Side(style=v) if v else Side()
    return Border(top=s(top), right=s(right), bottom=s(bottom), left=s(left))


def generate_excel(plate, car_model, date_from, date_to,
                   driver, odometer, refuel_set, agg, trip_purpose='dowóz/odbiór pracowników'):
    wb = Workbook()

    # ── Tytułowa ─────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Tytułowa'
    ws.sheet_view.showGridLines = False

    for col, w in [('A',4),('B',20),('C',30),('D',30),('E',30),('F',4)]:
        ws.column_dimensions[col].width = w

    # Header bar
    ws.merge_cells('B2:E2')
    c = ws['B2']
    c.value = f'Ewidencja Przebiegu Pojazdu nr. Rej.   {plate}'
    c.font  = Font(name='Arial', size=14, bold=True, color='000000')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 30  

    ws.merge_cells('B4:E4')
    c = ws['B4']
    c.value = 'wykorzystywanego wyłącznie do celów działalności gospodarczej'
    c.font  = Font(name='Arial', size=12, bold=True, color='000000')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 20

    # Thin dotted bottom border for fields
    def dotted_bdr():
        return Border(bottom=Side(style='dotted'))

    def info_row_centered(row, val, descr):
        ws.merge_cells(f'C{row}:D{row}')
        cv = ws[f'C{row}']
        cv.value = val
        cv.font = Font(name='Arial', size=11, color='000000')
        cv.alignment = Alignment(horizontal='center', vertical='bottom')
        cv.border = dotted_bdr()
        ws[f'D{row}'].border = dotted_bdr()
        
        ws.merge_cells(f'C{row+1}:D{row+1}')
        cd = ws[f'C{row+1}']
        cd.value = descr
        cd.font = Font(name='Arial', size=9, color='000000')
        cd.alignment = Alignment(horizontal='center', vertical='top')

    info_row_centered(6, 'Smart Work Sp. Z o.o.', 'Dane pracodawcy (nazwisko, imię/nazwa*)')
    ws.row_dimensions[6].height = 25
    info_row_centered(9, '58-100 Świdnica  Bystrzycka 7a', 'Adres prowadzonej działalności')
    ws.row_dimensions[9].height = 25
    info_row_centered(12, '8842754100', 'NIP')
    ws.row_dimensions[12].height = 25

    ws.row_dimensions[14].height = 20 # Odstęp
    
    def thick_bdr(top='thin', right='thin', bottom='thin', left='thin'):
        def s(v): return Side(style=v) if v else Side()
        return Border(top=s(top), right=s(right), bottom=s(bottom), left=s(left))

    total_km = round(sum(v['km'] for v in agg.values()), 2)
    odo_end  = (int(odometer) + int(total_km)) if odometer else 0

    # Rozpoczęcie
    ws.merge_cells('B15:E15')
    c = ws['B15']
    c.value = 'ROZPOCZĘCIE EWIDENCJI'
    c.font = Font(name='Arial', size=10, bold=True, color='000000')
    c.alignment = Alignment(horizontal='center', vertical='center')
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}15'].border = thick_bdr('medium', 'medium' if col=='E' else None, 'thin', 'medium' if col=='B' else None)
    ws.row_dimensions[15].height = 20
    
    ws.merge_cells('C16:E16')
    c = ws['C16']
    c.value = 'STAN LICZNIKA PRZEBIEGU POJAZDU\nna dzień rozpoczęcia prowadzenia ewidencji'
    c.font = Font(name='Arial', size=10, color='000000')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col in ['C', 'D', 'E']:
        ws[f'{col}16'].border = thick_bdr(None, 'medium' if col=='E' else None, 'thin', None)
    ws.row_dimensions[16].height = 30
    
    ws['B16'].value = 'DATA'
    ws['B16'].font = Font(name='Arial', size=10, color='000000')
    ws['B16'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B16'].border = thick_bdr(None, 'thin', 'thin', 'medium')
    
    ws['B17'].value = date_from
    ws['B17'].number_format = 'DD.MM.YYYY'
    ws['B17'].font = Font(name='Arial', size=14, bold=True, color='000000')
    ws['B17'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B17'].border = thick_bdr(None, 'thin', 'medium', 'medium')
    ws.row_dimensions[17].height = 25
    
    ws.merge_cells('C17:E17')
    b17 = ws['C17']
    b17.value = f"{int(odometer) if odometer else 0:,.0f} km".replace(',', ' ')
    b17.font = Font(name='Arial', size=14, bold=True, color='000000')
    b17.alignment = Alignment(horizontal='center', vertical='center')
    for col in ['C', 'D', 'E']:
        ws[f'{col}17'].border = thick_bdr(None, 'medium' if col=='E' else None, 'medium', None)
    
    ws.row_dimensions[18].height = 20 # Odstęp
    
    # Zakończenie
    ws.merge_cells('B19:E19')
    c = ws['B19']
    c.value = 'ZAKOŃCZENIE EWIDENCJI'
    c.font = Font(name='Arial', size=10, bold=True, color='000000')
    c.alignment = Alignment(horizontal='center', vertical='center')
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}19'].border = thick_bdr('medium', 'medium' if col=='E' else None, 'thin', 'medium' if col=='B' else None)
    ws.row_dimensions[19].height = 20
    
    ws.merge_cells('C20:D20')
    b20 = ws['C20']
    b20.value = 'STAN LICZNIKA PRZEBIEGU POJAZDU\nna dzień zakończenia prowadzenia ewidencji'
    b20.font = Font(name='Arial', size=9, color='000000')
    b20.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['C20'].border = thick_bdr(None, None, 'thin', None)
    ws['D20'].border = thick_bdr(None, 'thin', 'thin', None)
    ws.row_dimensions[20].height = 30
    
    ws['E20'].value = 'LICZBA PRZEJECHANYCH KILOMETRÓW\nna dzień zakończenia prowadzenia ewidencji'
    ws['E20'].font = Font(name='Arial', size=9, color='000000')
    ws['E20'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['E20'].border = thick_bdr(None, 'medium', 'thin', 'thin')
    
    ws['B20'].value = 'DATA'
    ws['B20'].font = Font(name='Arial', size=10, color='000000')
    ws['B20'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B20'].border = thick_bdr(None, 'thin', 'thin', 'medium')
    
    ws['B21'].value = date_to
    ws['B21'].number_format = 'DD.MM.YYYY'
    ws['B21'].font = Font(name='Arial', size=14, bold=True, color='000000')
    ws['B21'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B21'].border = thick_bdr(None, 'thin', 'medium', 'medium')
    ws.row_dimensions[21].height = 25
    
    ws.merge_cells('C21:D21')
    c21 = ws['C21']
    c21.value = f"{odo_end:,.0f} km".replace(',', ' ')
    c21.font = Font(name='Arial', size=14, bold=True, color='000000')
    c21.alignment = Alignment(horizontal='center', vertical='center')
    ws['C21'].border = thick_bdr(None, None, 'medium', None)
    ws['D21'].border = thick_bdr(None, 'thin', 'medium', None)
    
    ws['E21'].value = f"{total_km:.0f} km"
    ws['E21'].font = Font(name='Arial', size=14, bold=True, color='000000')
    ws['E21'].alignment = Alignment(horizontal='center', vertical='center')
    ws['E21'].border = thick_bdr(None, 'medium', 'medium', 'thin')

    ws.row_dimensions[22].height = 30 # Odstęp
    
    ws.merge_cells('C24:D24')
    cv = ws['C24']
    cv.value = (driver or '').strip()
    cv.font = Font(name='Arial', size=11, color='000000')
    cv.alignment = Alignment(horizontal='center', vertical='bottom')
    cv.border = dotted_bdr()
    ws['D24'].border = dotted_bdr()
    ws.row_dimensions[24].height = 20
    
    ws.merge_cells('C25:D25')
    cd = ws['C25']
    cd.value = 'podpis dysponenta'
    cd.font = Font(name='Arial', size=9, italic=True, color='666666')
    cd.alignment = Alignment(horizontal='center', vertical='top')

    # ── Rozlicznie ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Rozlicznie')
    ws2.sheet_view.showGridLines = False

    for col, w in [('A',5),('B',5),('C',13),('D',24),
                   ('E',42),('F',12),('G',24),('H',20)]:
        ws2.column_dimensions[col].width = w

    # Nagłówek nr rejestracyjny
    ws2.merge_cells('E4:G4')
    c = ws2['E4']
    c.value = 'Nr rejestracyjny auta: ................................................................'
    c.font = Font(name='Arial', size=9, bold=True, color='000000')
    c.alignment = Alignment(horizontal='right', vertical='bottom')
    
    c = ws2['H4']
    c.value = plate
    c.font = Font(name='Arial', size=9, bold=True, color='000000')
    c.alignment = Alignment(horizontal='right', vertical='bottom')
    
    def bdr_thick_outline(col, is_first, is_last, is_top=False, is_bottom=False):
        return Border(
            top=Side(style='medium' if is_top else 'thin'),
            bottom=Side(style='medium' if is_bottom else 'thin'),
            left=Side(style='medium' if is_first else 'thin'),
            right=Side(style='medium' if is_last else 'thin')
        )

    # Nagłówki kolumn
    HDRS = ['Nr\nkolejny\nwpisu', '', 'Data wyjazdu','Cel wyjazdu',
            'Opis trasy wyjazdu (skąd–dokąd)','Liczba\nfaktycznie\nprzejechanych\nkilometrów','Imię i nazwisko osoby kierującej pojazdem','Uwagi']
    
    for ci, h in enumerate(HDRS, 1):
        col = get_column_letter(ci); c = ws2[f'{col}6']
        c.value = h
        c.font  = Font(name='Arial', size=8, color='000000')
        c.fill  = PatternFill('solid', fgColor='D9D9D9') # Gray fill like in photo
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = bdr_thick_outline(ci, ci==1, ci==8, is_top=True)
    ws2.row_dimensions[6].height = 60

    # Dane dzienne
    year  = date_from.year  if date_from else datetime.date.today().year
    month = date_from.month if date_from else datetime.date.today().month
    days_in_month = calendar.monthrange(year, month)[1]
    ROW0 = 7

    running_odo = int(odometer) if odometer else 0

    mapped_refuel_set = {get_previous_working_day(rd) for rd in (refuel_set or set())}
    last_known_city = ""
    holidays = get_polish_holidays(year)

    for day in range(1, days_in_month + 1):
        d   = datetime.date(year, month, day)
        row = ROW0 + day - 1
        wd  = d.weekday()
        wknd = wd >= 5
        is_free = wknd or (d in holidays)
        bg   = 'FEE2E2' if is_free else ('FAFAFA' if day % 2 == 0 else 'FFFFFF')

        data  = agg.get(d)
        km    = round(data['km'], 2) if data and data['km'] > 0 else None
        
        route = ''
        if data and km:
            if data.get('daily_addresses') and len(data['daily_addresses']) > 1:
                routes_parts = []
                for daily_addrs in data['daily_addresses']:
                    rt, last_known_city = build_route(daily_addrs, last_known_city)
                    if rt:
                        routes_parts.append(rt)
                route = ' ;  '.join(routes_parts)
            else:
                route, last_known_city = build_route(data['addresses'], last_known_city)
                
        cel   = (trip_purpose or 'dowóz/odbiór pracowników') if km else ''
        uwagi = 'tankowanie' if d in mapped_refuel_set else ''

        if km:
            running_odo += km
            current_odo_display = running_odo
        else:
            current_odo_display = None

        row_data = [day, PL_DAYS[wd], d, cel, route, km if km is not None else 0, '', uwagi]

        for ci, val in enumerate(row_data, 1):
            col = get_column_letter(ci); c = ws2[f'{col}{row}']
            # In image, 0 km is explicitly shown 
            c.value  = val if val is not None else ''
            if ci == 6 and not km and not is_free:
                c.value = 0
                
            c.fill   = PatternFill('solid', fgColor='A6A6A6') if is_free else PatternFill('solid', fgColor='FFFFFF')
            c.border = bdr_thick_outline(ci, ci==1, ci==8)
            fnt_col  = '000000'

            if ci == 1:
                c.font = Font(name='Arial', size=9, color=fnt_col)
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 2:
                c.font = Font(name='Arial', size=9, color=fnt_col)
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 3:
                c.font = Font(name='Arial', size=9, color=fnt_col)
                if val: c.number_format = 'DD.MM.YYYY'
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 4:
                c.font = Font(name='Arial', size=9, color=fnt_col)
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 5:
                c.font = Font(name='Arial', size=9, color='000000')
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            elif ci == 6:
                c.font = Font(name='Arial', size=9, color='000000')
                if val == 0:
                    c.number_format = '0'
                elif val:
                    c.number_format = '#,##0'
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 7:
                if km:
                    c.value = (driver or '').strip()
                    c.font  = Font(name='Arial', size=9, color='000000')
                else:
                    c.font  = Font(name='Arial', size=9, color=fnt_col)
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 8:
                if uwagi:
                    c.font = Font(name='Arial', size=8, bold=True, color='B45309')
                    c.fill = PatternFill('solid', fgColor='FFFBEB')
                else:
                    c.font = Font(name='Arial', size=8, color=fnt_col)
                c.alignment = Alignment(horizontal='center', vertical='center', indent=1)

        # Removed setting explicit ws2.row_dimensions[row].height = 15 to allow auto-resizing.

    sr = ROW0 + days_in_month
    ws2.row_dimensions[sr].height = 20
    ws2.merge_cells(f'A{sr}:E{sr}')
    c = ws2[f'A{sr}']
    c.value = 'Koniec okresu rozliczeniowego'
    c.font  = Font(name='Arial', size=9, color='000000')
    c.fill  = PatternFill('solid', fgColor='D9D9D9')
    c.alignment = Alignment(horizontal='center', vertical='center')
    for ci in range(1, 6):
        col = get_column_letter(ci)
        ws2[f'{col}{sr}'].border = bdr_thick_outline(ci, ci==1, False, is_bottom=True)
    
    c_km = ws2[f'F{sr}']
    c_km.value = total_km; c_km.number_format = '#,##0'
    c_km.font  = Font(name='Arial', size=10, bold=True, color='000000')
    c_km.fill  = PatternFill('solid', fgColor='FFFFFF')
    c_km.alignment = Alignment(horizontal='center', vertical='center')
    c_km.border = bdr_thick_outline(6, False, False, is_bottom=True)
    
    # Przesuwamy puste tło i ramki dla wszystkich pozostałych kolumn az do H (indeks 8)
    ws2.merge_cells(f'G{sr}:H{sr}')
    ws2[f'G{sr}'].fill = PatternFill('solid', fgColor='D9D9D9')
    for ci in range(7, 9):
        col = get_column_letter(ci)
        ws2[f'{col}{sr}'].border = bdr_thick_outline(ci, False, ci==8, is_bottom=True)

    sig = sr + 2
    ws2.merge_cells(f'G{sig}:H{sig}')
    cv = ws2[f'G{sig}']
    cv.value = (driver or '').strip()
    cv.font  = Font(name='Arial', size=11, color='000000')
    cv.alignment = Alignment(horizontal='center', vertical='bottom')
    ws2[f'H{sig}'].font = Font(name='Arial', size=11, color='000000')
    
    ws2.merge_cells(f'G{sig+1}:H{sig+1}')
    cd = ws2[f'G{sig+1}']
    cd.value = 'podpis dysponenta'
    cd.font  = Font(name='Arial', size=8, italic=True, color=GRAY)
    cd.alignment = Alignment(horizontal='center', vertical='top')

    # Auto-adjust column widths based on content
    for col_idx, col in enumerate(ws2.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.row < 5 or cell.row >= sr or not getattr(cell, 'value', None):
                continue
            
            # If there are newlines, we find the longest line
            lines = str(cell.value).split('\n')
            for line in lines:
                if len(line) > max_length:
                    max_length = len(line)
                    
        if max_length > 0:
            if column_letter in ['A', 'B']:
                adjusted_width = max(5, min(max_length + 2, 8)) # Fixed width for Lp and Dz
            elif column_letter == 'E':
                adjusted_width = min(max_length + 2, 48)  # Limit column E width for A4 landscape
            else:
                adjusted_width = min(max_length + 2, 25)  # Limit other columns
            ws2.column_dimensions[column_letter].width = adjusted_width

    # Page Setup for A4 Landscape
    ws2.page_setup.orientation = ws2.ORIENTATION_LANDSCAPE
    ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
    ws2.page_setup.fitToPage = True
    ws2.page_setup.fitToHeight = 0  # Do not force height to 1 page
    ws2.page_setup.fitToWidth = 1   # Force width to 1 page
    ws2.print_options.horizontalCentered = True
    
    # Adjust margins to narrow so it fits better
    ws2.page_margins.left = 0.25
    ws2.page_margins.right = 0.25
    ws2.page_margins.top = 0.75
    ws2.page_margins.bottom = 0.75

    ws2.freeze_panes = f'A{ROW0}'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# FLASK ROUTES
# ══════════════════════════════════════════════════════════════════════════════

HTML_PAGE = r"""<!DOCTYPE html>
<html lang="pl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Ewidencja Przebiegu Pojazdu – Smart Work</title>
<style>
:root{--bg:#0B1120;--surface:#111827;--card:#1C2535;--border:#2D3748;--accent:#10B981;--accent2:#059669;--text:#F1F5F9;--muted:#94A3B8;--gray:#6B7280;--err:#EF4444;--radius:14px}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--text);font-family:'Segoe UI',system-ui,sans-serif;min-height:100vh;display:flex;align-items:flex-start;justify-content:center;padding:32px 16px 60px}
body::before{content:'';position:fixed;inset:0;background:radial-gradient(ellipse 700px 500px at 20% 10%,rgba(16,185,129,.07) 0%,transparent 60%),radial-gradient(ellipse 600px 400px at 80% 80%,rgba(16,185,129,.05) 0%,transparent 60%);pointer-events:none;z-index:0}
.wrap{width:100%;max-width:680px;position:relative;z-index:1;animation:up .5s ease}
@keyframes up{from{opacity:0;transform:translateY(18px)}to{opacity:1;transform:translateY(0)}}
.header{margin-bottom:32px}
.logo{display:flex;align-items:center;gap:14px;margin-bottom:14px}
.logo-icon{width:52px;height:52px;border-radius:14px;background:linear-gradient(135deg,var(--accent),#065F46);display:flex;align-items:center;justify-content:center;font-size:26px;box-shadow:0 0 28px rgba(16,185,129,.35);flex-shrink:0}
.logo-text{font-size:11px;letter-spacing:.2em;text-transform:uppercase;color:var(--accent);font-weight:600}
h1{font-size:26px;font-weight:800;line-height:1.25}
.subtitle{font-size:13px;color:var(--muted);margin-top:6px;line-height:1.6}
.card{background:var(--card);border:1px solid var(--border);border-radius:var(--radius);margin-bottom:16px;overflow:hidden;transition:border-color .2s}
.card:focus-within{border-color:rgba(16,185,129,.5)}
.card-head{padding:16px 22px 12px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:10px}
.card-head .icon{width:32px;height:32px;border-radius:8px;background:rgba(16,185,129,.12);display:flex;align-items:center;justify-content:center;font-size:16px;flex-shrink:0}
.card-head-text h3{font-size:13px;font-weight:600}
.card-head-text p{font-size:11px;color:var(--muted);margin-top:1px}
.card-body{padding:18px 22px 20px}
.upload-zone{border:2px dashed var(--border);border-radius:10px;padding:28px 20px;cursor:pointer;text-align:center;transition:all .22s;position:relative;background:var(--surface)}
.upload-zone input[type=file]{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%}
.upload-zone:hover,.upload-zone.drag{border-color:var(--accent);background:rgba(16,185,129,.04)}
.upload-zone.has-file{border-color:var(--accent);border-style:solid;background:rgba(16,185,129,.06)}
.upload-icon{font-size:32px;margin-bottom:8px;display:block}
.upload-title{font-size:14px;font-weight:600}
.upload-hint{font-size:11px;color:var(--muted);margin-top:4px;font-family:Consolas,monospace}
.upload-badge{display:none;margin-top:10px;background:rgba(16,185,129,.15);color:var(--accent);border:1px solid rgba(16,185,129,.3);border-radius:20px;padding:3px 10px;font-size:11px;font-family:Consolas,monospace}
.has-file .upload-badge{display:inline-block}
.has-file .upload-hint{color:var(--accent);opacity:.8}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:16px}
@media(max-width:500px){.grid2{grid-template-columns:1fr}}
.field{display:flex;flex-direction:column;gap:6px}
.field label{font-size:11px;font-weight:600;color:var(--muted);letter-spacing:.05em}
.field input{background:var(--surface);border:1.5px solid var(--border);border-radius:8px;padding:11px 14px;color:var(--text);font-family:inherit;font-size:13px;outline:none;transition:border-color .18s,box-shadow .18s}
.field input:focus{border-color:var(--accent);box-shadow:0 0 0 3px rgba(16,185,129,.12)}
.field input::placeholder{color:var(--muted)}
.refuel-wrap{display:flex;flex-wrap:wrap;gap:8px;margin-top:8px;min-height:32px}
.tag{display:flex;align-items:center;gap:6px;background:rgba(245,158,11,.12);border:1px solid rgba(245,158,11,.35);color:#FCD34D;border-radius:20px;padding:4px 10px 4px 12px;font-size:12px;animation:tagIn .2s ease}
@keyframes tagIn{from{opacity:0;transform:scale(.8)}to{opacity:1;transform:scale(1)}}
.tag .rm{cursor:pointer;opacity:.7;font-size:14px}
.tag .rm:hover{opacity:1;color:var(--err)}
.refuel-hint{font-size:11px;color:var(--muted);margin-top:6px}
.btn{width:100%;padding:16px;background:linear-gradient(135deg,var(--accent),var(--accent2));color:#052e16;font-size:15px;font-weight:800;letter-spacing:.03em;border:none;border-radius:var(--radius);cursor:pointer;transition:all .22s;display:flex;align-items:center;justify-content:center;gap:10px;position:relative;overflow:hidden}
.btn:hover{transform:translateY(-1px);box-shadow:0 10px 30px rgba(16,185,129,.3)}
.btn:active{transform:translateY(0)}
.btn:disabled{opacity:.5;cursor:not-allowed;transform:none;box-shadow:none}
.spinner{width:18px;height:18px;border:2.5px solid rgba(5,46,22,.3);border-top-color:#052e16;border-radius:50%;animation:spin .7s linear infinite;display:none}
@keyframes spin{to{transform:rotate(360deg)}}
.progress{height:4px;background:var(--border);border-radius:2px;margin-top:14px;overflow:hidden;display:none}
.progress-bar{height:100%;background:linear-gradient(90deg,var(--accent),#34D399);border-radius:2px;width:0%;transition:width .4s}
.progress.show{display:block}
.status{display:none;margin-top:14px;border-radius:10px;padding:14px 16px;font-size:13px;animation:up .3s ease;align-items:flex-start;gap:10px}
.status.ok{display:flex;background:rgba(16,185,129,.08);border:1px solid rgba(16,185,129,.3);color:var(--accent)}
.status.err{display:flex;background:rgba(239,68,68,.08);border:1px solid rgba(239,68,68,.3);color:#FCA5A5}
.status .si{font-size:18px;flex-shrink:0}
.info-box{display:flex;gap:10px;align-items:flex-start;background:rgba(255,255,255,.03);border:1px solid var(--border);border-radius:10px;padding:12px 16px;margin-top:16px;font-size:12px;color:var(--muted);line-height:1.6}
</style>
</head>
<body>
<div class="wrap">
  <div class="header">
    <div class="logo">
      <div class="logo-icon">🚗</div>
      <div>
        <div class="logo-text">Smart Work Sp. z o.o.</div>
      </div>
    </div>
    <h1>Generator Ewidencji<br>Przebiegu Pojazdu</h1>
    <p class="subtitle">Wgraj raport GPS → uzupełnij dane → pobierz gotowy plik Excel.<br>Weekendy automatycznie sumowane na piątek. Działa bez LibreOffice.</p>
  </div>

  <div class="card" id="gps-card">
    <div class="card-head">
      <div class="icon">📍</div>
      <div class="card-head-text">
        <h3>Plik GPS</h3>
        <p>Raport z trackera pojazdu (.XLS lub .XLSX)</p>
      </div>
    </div>
    <div class="card-body">
      <div class="upload-zone" id="zone">
        <input type="file" id="gps_file" accept=".xls,.xlsx" onchange="handleFile(this)">
        <span class="upload-icon">📂</span>
        <div class="upload-title" id="file-title">Kliknij lub przeciągnij plik tutaj</div>
        <div class="upload-hint" id="file-hint">.XLS · .XLSX · max 50 MB</div>
        <div class="upload-badge" id="file-badge">✓ plik wgrany</div>
      </div>
    </div>
  </div>

  <div class="card">
    <div class="card-head">
      <div class="icon">👤</div>
      <div class="card-head-text">
        <h3>Dane kierowcy</h3>
        <p>Imię i nazwisko pojawi się we wszystkich wierszach z km</p>
      </div>
    </div>
    <div class="card-body">
      <div class="grid2">
        <div class="field">
          <label>NAZWISKO I IMIĘ</label>
          <input type="text" id="driver_name" placeholder="np. Kowalski Jan" autocomplete="off">
        </div>
        <div class="field">
          <label>STAN LICZNIKA NA POCZĄTKU MIESIĄCA (KM)</label>
          <input type="number" id="odometer_start" placeholder="np. 248349" min="0">
        </div>
      </div>
      <div class="field" style="margin-top:14px;">
        <label>CEL WYJAZDU</label>
        <input type="text" id="trip_purpose" placeholder="np. dowóz/odbiór pracowników" autocomplete="off">
      </div>
    </div>
  </div>

  <div class="card" id="target-korekta-card">
    <div class="card-head">
      <div class="icon">🎯</div>
      <div class="card-head-text">
        <h3>Korekta końcowego przebiegu (Opcjonalnie)</h3>
        <p>Wpisz ręcznie docelowy stan licznika, aby zrównać trasy</p>
      </div>
    </div>
    <div class="card-body">
      <div class="grid2">
        <div class="field">
          <label>DOCELOWY STAN LICZNIKA KOŃCOWY</label>
          <input type="number" id="target_odometer" placeholder="np. 250000" min="0">
        </div>
        <div class="field" style="justify-content: flex-end; padding-bottom: 8px;">
          <label style="display:flex; align-items:center; gap:8px; cursor:pointer; font-size:13px; color:var(--text); font-weight:600;">
            <input type="checkbox" id="adjust_mileage" checked style="width:18px; height:18px; accent-color: var(--accent);">
            Podciągnij trasy, aby zrównać
          </label>
        </div>
      </div>
    </div>
  </div>

  <div class="card">
    <div class="card-head">
      <div class="icon">⛽</div>
      <div class="card-head-text">
        <h3>Dni tankowania</h3>
        <p>Wpisz daty – pojawią się w kolumnie Uwagi jako "tankowanie"</p>
      </div>
    </div>
    <div class="card-body">
      <div class="field">
        <label>WPISZ DATĘ I NACIŚNIJ ENTER LUB SPACJĘ</label>
        <input type="text" id="refuel_input" placeholder="np. 5.12 lub 05.12.2025" autocomplete="off">
      </div>
      <div class="refuel-wrap" id="refuel-tags"></div>
      <div class="refuel-hint">📌 Format: DD.MM lub DD.MM.RRRR &nbsp;|&nbsp; Możesz dodać wiele dat</div>
    </div>
  </div>

  <div class="card">
    <div class="card-head">
      <div class="icon">📅</div>
      <div class="card-head-text">
        <h3>Tryb zapisu przebiegu</h3>
        <p>Wybierz jak mają być przypisane kilometry w ewidencji</p>
      </div>
    </div>
    <div class="card-body">
      <label style="display:flex;align-items:center;gap:10px;cursor:pointer;font-size:13px;color:var(--text);font-weight:600;">
        <input type="checkbox" id="use_actual_days" style="width:18px;height:18px;accent-color:var(--accent);">
        Zapisz przebieg w faktycznych dniach jazdy (bez przesuwania na piątek)
      </label>
      <p style="font-size:11px;color:var(--muted);margin-top:8px;line-height:1.6;">
        <strong style="color:var(--text);">Domyślnie (odznaczone):</strong> km z weekendu/świąt → poprzedni piątek roboczy.<br>
        <strong style="color:var(--text);">Po zaznaczeniu:</strong> km zapisane dokładnie w dniu, w którym jeździło auto.
      </p>
    </div>
  </div>

  <button class="btn" id="btn" onclick="generate()">

    <div class="spinner" id="spinner"></div>
    <span id="btn-text">⚡ Generuj ewidencję Excel</span>
  </button>
  <div class="progress" id="progress"><div class="progress-bar" id="pbar"></div></div>
  <div class="status" id="status"><div class="si" id="s-icon"></div><div id="s-msg"></div></div>

  <div class="info-box">
    <div style="font-size:16px;flex-shrink:0">💡</div>
    <div>Aplikacja odczytuje dane GPS, sumuje kilometry per dzień i tworzy plik Excel ze stroną tytułową i arkuszem rozliczeniowym. Dane z soboty i niedzieli przenoszone są automatycznie na poprzedni piątek. <strong style="color:#10B981">Nie wymaga LibreOffice.</strong></div>
  </div>
</div>
<script>
const refuelDates=[];
const zone=document.getElementById('zone');
zone.addEventListener('dragover',e=>{e.preventDefault();zone.classList.add('drag')});
zone.addEventListener('dragleave',()=>zone.classList.remove('drag'));
zone.addEventListener('drop',e=>{e.preventDefault();zone.classList.remove('drag');if(e.dataTransfer.files.length){document.getElementById('gps_file').files=e.dataTransfer.files;handleFile(document.getElementById('gps_file'))}});
function handleFile(input){if(input.files&&input.files[0]){const name=input.files[0].name;zone.classList.add('has-file');document.getElementById('file-title').textContent=name;document.getElementById('file-hint').textContent=(input.files[0].size/1024/1024).toFixed(2)+' MB';document.getElementById('file-badge').textContent='✓ '+name;document.getElementById('gps-card').style.borderColor='rgba(16,185,129,.6)'}}
document.getElementById('refuel_input').addEventListener('keydown',e=>{if(e.key==='Enter'||e.key===' '||e.key===','){e.preventDefault();const v=e.target.value.trim();if(v){addTag(v);e.target.value=''}}});
document.getElementById('refuel_input').addEventListener('blur',e=>{const v=e.target.value.trim();if(v){addTag(v);e.target.value=''}});
function addTag(val){val=val.replace(/[,;]/g,'').trim();if(!val||refuelDates.includes(val))return;if(!/^\d{1,2}[.\-\/]\d{1,2}([.\-\/]\d{2,4})?$/.test(val)){showStatus('err','⚠','Nieprawidłowy format: '+val+'. Użyj DD.MM');return}refuelDates.push(val);const wrap=document.getElementById('refuel-tags');const tag=document.createElement('div');tag.className='tag';tag.innerHTML='⛽ '+val+' <span class="rm" data-val="'+val+'">✕</span>';tag.querySelector('.rm').onclick=function(){refuelDates.splice(refuelDates.indexOf(this.dataset.val),1);this.parentElement.remove()};wrap.appendChild(tag)}
let progIv=null;
function startProgress(){const wrap=document.getElementById('progress');const bar=document.getElementById('pbar');wrap.classList.add('show');bar.style.width='0%';let w=0;progIv=setInterval(()=>{w=Math.min(w+Math.random()*7,88);bar.style.width=w+'%'},200)}
function stopProgress(){clearInterval(progIv);document.getElementById('pbar').style.width='100%';setTimeout(()=>document.getElementById('progress').classList.remove('show'),500)}
function showStatus(type,icon,msg){const el=document.getElementById('status');el.className='status '+type;document.getElementById('s-icon').textContent=icon;document.getElementById('s-msg').textContent=msg}
async function generate(){
  const file=document.getElementById('gps_file').files[0];
  if(!file){showStatus('err','⚠','Wybierz plik GPS.');return}
  const btn=document.getElementById('btn');const spin=document.getElementById('spinner');const btxt=document.getElementById('btn-text');
  btn.disabled=true;spin.style.display='block';btxt.textContent='Przetwarzanie...';
  document.getElementById('status').className='status';
  startProgress();
  try{
    const form=new FormData();
    form.append('gps_file',file);
    form.append('driver_name',document.getElementById('driver_name').value);
    form.append('odometer_start',document.getElementById('odometer_start').value||'0');
    form.append('refuel_dates',refuelDates.join(','));
    form.append('target_odometer',document.getElementById('target_odometer').value||'0');
    form.append('adjust_mileage',document.getElementById('adjust_mileage').checked);
    form.append('use_actual_days',document.getElementById('use_actual_days').checked);
    form.append('trip_purpose',document.getElementById('trip_purpose').value.trim());
    const resp=await fetch('/generate',{method:'POST',body:form});
    stopProgress();
    if(!resp.ok){const data=await resp.json().catch(()=>({}));throw new Error(data.error||'Błąd serwera ('+resp.status+')')}
    const blob=await resp.blob();
    const cd=resp.headers.get('Content-Disposition')||'';
    const m=cd.match(/filename[^=\n]*=(["']?)(.+?)\1/);
    const name=m?m[2]:'ewidencja.xlsx';
    const url=URL.createObjectURL(blob);const a=document.createElement('a');a.href=url;a.download=name;document.body.appendChild(a);a.click();document.body.removeChild(a);URL.revokeObjectURL(url);
    showStatus('ok','✓','Plik "'+name+'" pobrany!');
  }catch(err){stopProgress();showStatus('err','✗','Błąd: '+err.message)}
  finally{btn.disabled=false;spin.style.display='none';btxt.textContent='⚡ Generuj ewidencję Excel'}
}
</script>
</body>
</html>"""


@app.route('/')
def index():
    resp = make_response(HTML_PAGE)
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    return resp


@app.route('/generate', methods=['POST'])
def generate():
    try:
        if 'gps_file' not in request.files:
            return jsonify({'error': 'Brak pliku GPS'}), 400

        gps_file     = request.files['gps_file']
        driver       = request.form.get('driver_name', '').strip()
        odo_raw      = request.form.get('odometer_start', '0').strip()
        refuel_raw   = request.form.get('refuel_dates', '').strip()
        target_odo_raw = request.form.get('target_odometer', '').strip()
        adjust_mil     = request.form.get('adjust_mileage', 'false') == 'true'
        use_actual     = request.form.get('use_actual_days', 'false') == 'true'
        trip_purpose   = request.form.get('trip_purpose', '').strip()

        try:
            odometer = int(odo_raw.replace(' ', '').replace(',', '')) if odo_raw else 0
        except ValueError:
            odometer = 0

        try:
            target_odometer = int(target_odo_raw.replace(' ', '').replace(',', '')) if target_odo_raw else 0
        except ValueError:
            target_odometer = 0

        file_bytes = gps_file.read()
        plate, car_model, date_from, date_to, day_groups = \
            parse_gps(file_bytes, gps_file.filename)

        if not day_groups:
            return jsonify({'error': 'Nie znaleziono danych GPS w pliku.'}), 400

        agg = aggregate_actual(day_groups) if use_actual else aggregate(day_groups)

        # Odometer Proportional Adjustment
        if adjust_mil and target_odometer > odometer:
            expected_gps_km = float(target_odometer - odometer)
            current_gps_km = sum(v['km'] for v in agg.values())
            
            if current_gps_km > 0 and abs(expected_gps_km - current_gps_km) > 0.01:
                ratio = expected_gps_km / current_gps_km
                for k in agg:
                    if agg[k]['km'] > 0:
                        agg[k]['km'] = round(agg[k]['km'] * ratio, 2)
                
                # Fix rounding diffs
                new_total = sum(v['km'] for v in agg.values())
                diff = round(expected_gps_km - new_total, 2)
                if abs(diff) > 0:
                    max_day = max((k for k in agg if agg[k]['km'] > 0), key=lambda x: agg[x]['km'], default=None)
                    if max_day:
                        agg[max_day]['km'] = round(agg[max_day]['km'] + diff, 2)

        # Parsuj daty tankowania
        year  = date_from.year  if date_from else datetime.date.today().year
        month = date_from.month if date_from else datetime.date.today().month
        refuel_set = set()
        for part in re.split(r'[,;\s]+', refuel_raw):
            part = part.strip()
            if not part: continue
            m = re.match(r'(\d{1,2})[.\-/](\d{1,2})(?:[.\-/](\d{2,4}))?', part)
            if m:
                try:
                    d  = int(m.group(1))
                    mo = int(m.group(2))
                    y  = int(m.group(3)) if m.group(3) else year
                    if y < 100: y += 2000
                    refuel_set.add(datetime.date(y, mo, d))
                except Exception:
                    pass

        buf = generate_excel(plate, car_model, date_from, date_to,
                             driver, odometer, refuel_set, agg,
                             trip_purpose=trip_purpose)

        # Odtworzenie nazwy z przesłanego pliku i wymuszenie .xlsx
        original_filename = gps_file.filename
        if original_filename:
            base_name = os.path.splitext(original_filename)[0]
            filename = f"{base_name}.xlsx"
        else:
            filename = 'ewidencja.xlsx'
        
        resp = send_file(buf, as_attachment=True,
                         download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # Zapewnia widoczność nazwy pliku w nagłówku dla fetch API w przeglądarce
        resp.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return resp

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        return jsonify({'error': str(e), 'detail': tb}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
